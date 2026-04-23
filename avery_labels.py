#!/usr/bin/env python3
"""
Avery Label Generator - Interactive CLI
Generate labels from CSV/Excel/JSON using Avery sheet specs

Usage (guided mode):
    python avery_labels.py -i data.xlsx

Usage (arguments):
    python avery_labels.py -i data.xlsx -o labels.pdf --spec 5960 -f "Store Name" -f "confirmation #"
    
List options:
    python avery_labels.py --list-specs
    python avery_labels.py -i data.xlsx --list
"""

import argparse
import csv
import json
import math
import os
import sys
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from reportlab.lib import pagesizes
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader
except ImportError:
    print("ERROR: reportlab not installed. Run: pip install reportlab openpyxl")
    sys.exit(1)

DEFAULT_FONT = "Helvetica"
DEFAULT_FONT_SIZE = 10
SPECS_FILE = "avery_specs.csv"


def load_specs_from_csv(csv_path: str = None) -> dict:
    """Load Avery specs from CSV file."""
    if csv_path is None:
        csv_path = Path(__file__).parent / SPECS_FILE
    
    specs = {}
    
    if not Path(csv_path).exists():
        print(f"ERROR: Specs file not found: {csv_path}")
        sys.exit(1)
    
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get("labelTypeName", "")
            if not name:
                continue
            
            spec_key = name.replace("Avery ", "").strip()
            
            specs[spec_key] = {
                "label_width": float(row.get("labelSheetWidth", 8.5)) * inch,
                "label_height": (
                    (float(row.get("labelSheetHeight", 11)) - 
                     float(row.get("topMargin", 0.5)) - 
                     float(row.get("bottomMargin", 0.5)))
                    / float(row.get("numberOfRows", 10))
                ) * inch,
                "vertical_pitch": (
                    (float(row.get("labelSheetHeight", 11)) - 
                     float(row.get("topMargin", 0.5)) - 
                     float(row.get("bottomMargin", 0.5)))
                    / float(row.get("numberOfRows", 10))
                ) * inch,
                "columns": int(row.get("numberOfColumns", 3)),
                "rows": int(row.get("numberOfRows", 10)),
                "page_width": float(row.get("labelSheetWidth", 8.5)) * inch,
                "page_height": float(row.get("labelSheetHeight", 11)) * inch,
                "margin_left": float(row.get("leftMargin", 0.188)) * inch,
                "margin_top": float(row.get("topMargin", 0.5)) * inch,
                "h_gap": float(row.get("horizontalGutter", 0.125)) * inch,
                "v_gap": float(row.get("verticalGutter", 0)) * inch,
            }
    
    return specs


def list_specs():
    """List available Avery specs."""
    specs = load_specs_from_csv()
    print("\nAvailable Avery label formats:")
    print("-" * 50)
    for name, spec in sorted(specs.items()):
        print(f"  {name:12}  {spec['columns']} cols x {spec['rows']:2} rows = {spec['columns'] * spec['rows']:2} labels/sheet")
    print("-" * 50)
    return specs


def list_columns(filepath: str):
    """List available columns in input file."""
    ext = Path(filepath).suffix.lower()
    columns = []
    
    if ext == ".csv":
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            columns = reader.fieldnames or []
    elif ext in [".xlsx", ".xls"]:
        if openpyxl is None:
            print("ERROR: openpyxl not installed")
            sys.exit(1)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        columns = [cell.value for cell in ws[1]]
    elif ext == ".json":
        with open(filepath, encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list) and len(data) > 0:
            columns = list(data[0].keys())
    else:
        print(f"ERROR: Unsupported format: {ext}")
        sys.exit(1)
    
    print(f"\nAvailable columns in '{Path(filepath).name}':")
    print("-" * 50)
    for i, col in enumerate(columns, 1):
        print(f"  {i}. {col}")
    print("-" * 50)
    return columns


def ask_input(prompt: str, default: str = None, required: bool = True) -> str:
    """Interactive prompt with optional default."""
    if default:
        response = input(f"{prompt} [{default}]: ").strip()
        return response if response else default
    else:
        if required:
            while True:
                response = input(f"{prompt}: ").strip()
                if response:
                    return response
        else:
            response = input(f"{prompt} (optional): ").strip()
            return response


def ask_choice(prompt: str, options: list, default: int = None) -> str:
    """Ask user to choose from a list of numbered options."""
    print(f"\n{prompt}:")
    for i, opt in enumerate(options, 1):
        print(f"  {i}. {opt}")
    
    if default:
        prompt = f"\n{prompt} [{default}]: "
    else:
        prompt = "\nEnter number: "
    
    while True:
        try:
            response = input(prompt).strip()
            if not response and default:
                return str(default)
            idx = int(response) - 1
            if 0 <= idx < len(options):
                return options[idx]
        except ValueError:
            pass
        print("Invalid choice. Try again.")


def ask_multichoice(prompt: str, options: list) -> list:
    """Ask user to select multiple options (comma-separated)."""
    print(f"\n{prompt}:")
    for i, opt in enumerate(options, 1):
        print(f"  {i}. {opt}")
    print("\nEnter numbers separated by commas (e.g., 1,3,5) or press Enter to skip:")
    
    while True:
        response = input(": ").strip()
        if not response:
            return []
        
        try:
            indices = [int(x.strip()) - 1 for x in response.split(",")]
            selected = [options[i] for i in indices if 0 <= i < len(options)]
            if selected:
                print(f"  Selected: {', '.join(selected)}")
                return selected
        except (ValueError, IndexError):
            pass
        print("Invalid. Try again.")


def read_csv(filepath: str) -> list[dict]:
    """Read CSV file and return list of row dicts."""
    rows = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def read_excel(filepath: str) -> list[dict]:
    """Read Excel file and return list of row dicts."""
    if openpyxl is None:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    headers = [h.strip() if isinstance(h, str) else f"col_{i}" for i, h in enumerate(headers)]
    
    rows = []
    for row in ws.iter_rows(min_row=2):
        row_data = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = cell.value
        rows.append(row_data)
    
    return rows


def read_json(filepath: str) -> list[dict]:
    """Read JSON file and return list of row dicts."""
    with open(filepath, encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data
    return [data]


def read_input(filepath: str) -> list[dict]:
    """Auto-detect format and read input file."""
    ext = Path(filepath).suffix.lower()
    
    if ext == ".csv":
        return read_csv(filepath)
    elif ext in [".xlsx", ".xls"]:
        return read_excel(filepath)
    elif ext == ".json":
        return read_json(filepath)
    else:
        print(f"ERROR: Unsupported file format: {ext}")
        sys.exit(1)


def format_label_text(row: dict, fields: list[str], separator: str = "\n") -> str:
    """Format label text from row data using specified fields."""
    parts = []
    
    for field in fields:
        if field in row:
            val = row.get(field, "")
            if val is not None and str(val).strip() != "":
                parts.append(str(val).strip())
    
    return separator.join(parts)


def draw_label(c: canvas.Canvas, x: float, y: float, specs: dict, text: str, font_size: int = None):
    """Draw a single label at position (x, y) from bottom-left (PDF coords)."""
    width = specs["label_width"]
    height = specs["label_height"]
    
    c.saveState()
    c.setFont(DEFAULT_FONT, font_size or DEFAULT_FONT_SIZE)
    
    lines = text.split("\n")
    text_y = y + height - 0.25 * inch
    
    for line in lines:
        c.drawString(x + 0.125 * inch, text_y, line)
        text_y -= (font_size or DEFAULT_FONT_SIZE) * 1.2
    
    c.restoreState()


def generate_labels(
    data: list[dict],
    output: str,
    specs: dict,
    fields: list[str],
    mode: str = "unique",
    separator: str = "\n",
):
    """Generate PDF with labels."""
    
    c = canvas.Canvas(output, pagesize=(specs["page_width"], specs["page_height"]))
    
    margin_left = specs["margin_left"]
    margin_top = specs["margin_top"]
    label_width = specs["label_width"]
    vertical_pitch = specs.get("vertical_pitch", specs["label_height"])
    h_gap = specs["h_gap"]
    columns = specs["columns"]
    rows = specs["rows"]
    
    if mode == "repeat":
        if not data:
            print("ERROR: No data to repeat")
            sys.exit(1)
        
        label_text = format_label_text(data[0], fields, separator)
        
        for row_idx in range(rows):
            for col_idx in range(columns):
                x = margin_left + col_idx * (label_width + h_gap)
                y = margin_top + (rows - 1 - row_idx) * vertical_pitch
                draw_label(c, x, y, specs, label_text)
        
        c.save()
        print(f"\nCreated {output} - repeated label ({columns * rows} per page)")
        return
    
    label_count = 0
    current_row = 0
    current_col = 0
    
    for row_data in data:
        label_text = format_label_text(row_data, fields, separator)
        if not label_text.strip():
            continue
        
        x = margin_left + current_col * (label_width + h_gap)
        y = margin_top + (rows - 1 - current_row) * vertical_pitch
        
        draw_label(c, x, y, specs, label_text)
        label_count += 1
        
        current_col += 1
        if current_col >= columns:
            current_col = 0
            current_row += 1
        
        if current_row >= rows and label_count < len(data):
            c.showPage()
            current_row = 0
            current_col = 0
    
    c.save()
    pages = math.ceil(label_count / (columns * rows)) if label_count > 0 else 1
    print(f"\n✓ Created {output} - {label_count} labels ({pages} pages)")


def guided_mode(input_file: str = None):
    """Interactive guided mode."""
    print("\n" + "="*50)
    print("  AVERY LABEL GENERATOR - Guided Mode")
    print("="*50)
    
    # Step 1: Input file
    if not input_file:
        input_file = ask_input("Input file (CSV, XLSX, JSON)")
    
    if not Path(input_file).exists():
        print(f"ERROR: File not found: {input_file}")
        sys.exit(1)
    
    print(f"\n[1] Loading {input_file}...")
    data = read_input(input_file)
    print(f"    Found {len(data)} rows")
    
    # Step 2: List columns and let user select
    columns = list_columns(input_file)
    if not columns:
        print("ERROR: No columns found in file")
        sys.exit(1)
    
    selected_fields = ask_multichoice("Select fields to include on label", columns)
    if not selected_fields:
        print("ERROR: Must select at least one field")
        sys.exit(1)
    
    # Step 3: Label format
    specs = load_specs_from_csv()
    spec_names = sorted(specs.keys())
    default_spec = "5960"
    default_idx = spec_names.index(default_spec) + 1 if default_spec in spec_names else 1
    
    spec_name = ask_choice("Select label format", spec_names, default_idx)
    
    # Step 4: Output file
    default_output = input_file.rsplit(".", 1)[0] + "_labels.pdf"
    output_file = ask_input("Output filename", default_output)
    
    # Step 5: Mode
    print("\n[5] Mode:")
    print("  1. Unique - each row = one label")
    print("  2. Repeat - same label on all slots")
    mode_choice = ask_input("Mode", "1")
    mode = "unique" if mode_choice == "1" else "repeat"
    
    print("\n[6] Generating labels...")
    generate_labels(
        data=data,
        output=output_file,
        specs=specs[spec_name],
        fields=selected_fields,
        mode=mode,
    )
    
    print(f"\nDone! Output: {output_file}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Avery labels from CSV/Excel/JSON",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    
    parser.add_argument("--input", "-i", help="Input file (CSV, XLSX, JSON)")
    parser.add_argument("--output", "-o", default="labels.pdf", help="Output PDF file")
    parser.add_argument("--list", action="store_true", help="List columns in input file")
    parser.add_argument("--list-specs", action="store_true", help="List available Avery specs")
    parser.add_argument("--spec", help="Avery spec name (e.g., 5960)")
    parser.add_argument("--field", "-f", action="append", dest="fields", help="Field/column to include")
    parser.add_argument("--separator", default="\n", help="Field separator")
    parser.add_argument("--mode", choices=["unique", "repeat"], default="unique")
    
    args = parser.parse_args()
    
    # List-only modes
    if args.list_specs:
        list_specs()
        return
    
    if args.list:
        if not args.input:
            print("ERROR: --list requires --input")
            sys.exit(1)
        list_columns(args.input)
        return
    
    # Guided mode if no input
    if not args.input:
        print("="*50)
        print("  AVERY LABEL GENERATOR")
        print("="*50)
        print("Usage: python avery_labels.py -i data.xlsx")
        print("Run with -i to start guided mode or see --help for arguments")
        print("="*50)
        guided_mode()
        return
    
    # Require at least one field
    if not args.fields:
        columns = list_columns(args.input)
        print("\nWhich column(s) to use? Enter numbers (e.g., 1,3): ")
        response = input(": ").strip()
        try:
            indices = [int(x.strip()) - 1 for x in response.split(",")]
            args.fields = [columns[i] for i in indices if 0 <= i < len(columns)]
        except (ValueError, IndexError):
            pass
        
        if not args.fields:
            print("ERROR: Must specify at least one field with --field")
            sys.exit(1)
    
    # Load spec
    specs = load_specs_from_csv()
    if args.spec:
        if args.spec not in specs:
            print(f"ERROR: Unknown spec '{args.spec}'")
            print("Use --list-specs to see available")
            sys.exit(1)
        spec = specs[args.spec]
    else:
        # Ask for spec
        spec_name = ask_choice("Select label format", sorted(specs.keys()), 1)
        spec = specs[spec_name]
    
    # Read data
    print(f"Reading {args.input}...")
    data = read_input(args.input)
    print(f"Found {len(data)} rows")
    
    # Generate
    generate_labels(
        data=data,
        output=args.output,
        specs=spec,
        fields=args.fields,
        mode=args.mode,
        separator=args.separator,
    )


if __name__ == "__main__":
    main()